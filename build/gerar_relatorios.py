#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Auxiliar da aba RELATÓRIOS — calcula as MÉTRICAS por período do funil.

Uso pela Routine diária (1×/dia às 23h59 BRT) que regenera os briefings do Gestor:
este script NÃO escreve texto e NÃO chama nenhuma IA/API. Ele só faz a
matemática (fonte única de verdade, reaproveitando build.process) e emite um
JSON com os números de cada período. A Routine (Claude Code, via assinatura,
sem consumir créditos da API) lê esses números + build/GUIA-RELATORIOS.md e
redige os 9 briefings em build/relatorios.json.

Entradas (uma das duas):
    --meta-file meta.csv --sales-file sales.csv    (CSVs já extraídos)
    --xlsx central.xlsx                            (workbook com as abas
                                                    'Meta Ads…' e 'Compras…')

Saída:
    --out relatorios_metrics.json   (default: build/relatorios_metrics.json)

Exemplos:
    python build/gerar_relatorios.py --meta-file meta.csv --sales-file sales.csv
    python build/gerar_relatorios.py --xlsx central.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import date, datetime, timedelta, timezone

# Reaproveita TODA a lógica de leitura/atribuição do build (mesma fonte de verdade)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build  # noqa: E402

BRT = timezone(timedelta(hours=-3))


# --------------------------------------------------------------------------- #
# Extração das abas a partir de um .xlsx (opcional)
# --------------------------------------------------------------------------- #
def extract_xlsx(xlsx_path: str, meta_out: str, sales_out: str) -> None:
    import openpyxl  # dependência só neste caminho
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def dump(sheet_name: str, out: str) -> None:
        ws = wb[sheet_name]
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if c is None else c for c in row])

    meta_sheet = next((s for s in wb.sheetnames if s.lower().startswith("meta ads")), None)
    sales_sheet = next((s for s in wb.sheetnames if s.lower().startswith("compras")
                        or s.lower().startswith("compradores")), None)
    if not meta_sheet or not sales_sheet:
        raise SystemExit(f"Abas não encontradas. Disponíveis: {wb.sheetnames}")
    dump(meta_sheet, meta_out)
    dump(sales_sheet, sales_out)


# --------------------------------------------------------------------------- #
# Métricas (mesma matemática do app.js: derive/agg)
# --------------------------------------------------------------------------- #
def new_bucket() -> dict:
    return dict(sp=0.0, im=0.0, cl=0.0, pv=0.0, ck=0.0, vendas=0, vendasM=0, fat=0.0)


def add_meta(a: dict, m: dict) -> None:
    for k in ("sp", "im", "cl", "pv", "ck"):
        a[k] += m[k]


def add_sale(a: dict, s: dict) -> None:
    a["vendas"] += s["main"]
    a["vendasM"] += 1 if (s["main"] and s["meta"]) else 0
    a["fat"] += s["val"]


def sd(n, d):
    return (n / d) if d else None


def derive(a: dict, tax: float) -> dict:
    g = a["sp"] * tax
    return dict(
        gasto=g, impr=a["im"], cliques=a["cl"], pv=a["pv"], ck=a["ck"], vendas=a["vendas"], fat=a["fat"],
        cpm=sd(g * 1000, a["im"]), ctr=sd(a["cl"], a["im"]), cpc=sd(g, a["cl"]), cpv=sd(g, a["pv"]),
        cr=sd(a["pv"], a["cl"]), cpic=sd(g, a["ck"]), vischk=sd(a["ck"], a["pv"]),
        convchk=sd(a["vendasM"], a["ck"]), convlp=sd(a["vendasM"], a["pv"]),
        cac=sd(g, a["vendas"]), roas=sd(a["fat"], g), ticket=sd(a["fat"], a["vendas"]),
    )


def in_range(d, a, b) -> bool:
    return bool(d) and a <= d <= b


def presets(today: str, dmin: str, dmax: str) -> dict:
    def addd(s, n):
        return (date.fromisoformat(s) + timedelta(days=n)).isoformat()

    def month_first(s):
        return s[:8] + "01"

    y, m, _ = today.split("-")
    y, m = int(y), int(m)
    last_day_prev = date(y, m, 1) - timedelta(days=1)
    first_prev = last_day_prev.replace(day=1)
    return {
        "hoje": (today, today),
        "ontem": (addd(today, -1), addd(today, -1)),
        "3d": (addd(today, -2), today),
        "7d": (addd(today, -6), today),
        "14d": (addd(today, -13), today),
        "30d": (addd(today, -29), today),
        "mes": (month_first(today), today),
        "mespass": (first_prev.isoformat(), last_day_prev.isoformat()),
        "todo": (dmin, dmax),
    }


def totals(meta, sales, a, b, ads_only):
    bk = new_bucket()
    for m in meta:
        if in_range(m["d"], a, b):
            add_meta(bk, m)
    for s in sales:
        if in_range(s["d"], a, b):
            if ads_only and not s["meta"]:
                continue
            add_sale(bk, s)
    return bk


def by_dim(meta, sales, a, b, dim):
    mp = {}
    for m in meta:
        if in_range(m["d"], a, b):
            mp.setdefault(m[dim], new_bucket())
            add_meta(mp[m[dim]], m)
    for s in sales:
        if in_range(s["d"], a, b) and s["meta"]:
            mp.setdefault(s[dim], new_bucket())
            add_sale(mp[s[dim]], s)
    return mp


def r(v, p=4):
    return None if v is None else round(v, p)


def period_metrics(meta, sales, a, b, tax):
    dT, dA = derive(totals(meta, sales, a, b, False), tax), derive(totals(meta, sales, a, b, True), tax)

    def pack(d):
        return {k: r(d[k], 2 if k in ("gasto", "cpm", "cpc", "cpv", "cpic", "cac", "fat", "ticket", "roas") else 4)
                for k in d}

    camps = []
    for name, bk in sorted(by_dim(meta, sales, a, b, "camp").items(), key=lambda kv: -kv[1]["sp"]):
        d = derive(bk, tax)
        camps.append({"nome": name, **pack(d)})
    ads = []
    for name, bk in by_dim(meta, sales, a, b, "ad").items():
        if bk["sp"] <= 0:
            continue
        d = derive(bk, tax)
        ads.append({"nome": name, "gasto": r(d["gasto"], 2), "vendas": d["vendas"],
                    "cac": r(d["cac"], 2), "roas": r(d["roas"], 2)})
    ads.sort(key=lambda x: (-(x["roas"] if x["roas"] is not None else -1),
                            x["cac"] if x["cac"] is not None else 9e9))
    top = ads[:5]
    top_names = {x["nome"] for x in top}
    worst = [x for x in ads if x["nome"] not in top_names][-5:][::-1]
    return {"de": a, "ate": b, "total": pack(dT), "ads": pack(dA),
            "campanhas": camps, "top_anuncios": top, "piores_anuncios": worst}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--meta-file")
    ap.add_argument("--sales-file")
    ap.add_argument("--xlsx", help="workbook .xlsx com as abas Meta Ads/Compras")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                                  "relatorios_metrics.json"))
    args = ap.parse_args()

    meta_file, sales_file = args.meta_file, args.sales_file
    if args.xlsx:
        base = os.path.dirname(os.path.abspath(args.out))
        meta_file = os.path.join(base, "_meta_tmp.csv")
        sales_file = os.path.join(base, "_sales_tmp.csv")
        extract_xlsx(args.xlsx, meta_file, sales_file)

    # Sem --meta-file/--sales-file/--xlsx: busca ao vivo via CSV público do
    # Google Sheets (mesmo caminho do build.py) — usado pelo GitHub Actions,
    # que alcança docs.google.com (o sandbox do agente não alcança).
    meta_rows = build.load_rows(build.META_CSV_URL, meta_file)
    sales_rows = build.load_rows(build.SALES_CSV_URL, sales_file)
    data = build.process(meta_rows, sales_rows)
    meta, sales, B = data["meta"], data["sales"], data["build"]
    tax = B["tax_factor"]  # imposto Meta aplicado (mesmo default da UI: ON)

    ps = presets(B["today"], B["date_min"], B["date_max"])
    out = {
        "gerado_em": datetime.now(BRT).strftime("%d/%m/%Y %H:%M"),
        "hoje": B["today"], "periodo_dados": {"de": B["date_min"], "ate": B["date_max"]},
        "metas": {"cac": B.get("cac_target"), "roas": B.get("roas_target")},
        "ticket_medio_geral": r(derive(totals(meta, sales, B["date_min"], B["date_max"], False), tax)["ticket"], 2),
        "periodos": {k: period_metrics(meta, sales, a, b, tax) for k, (a, b) in ps.items()},
    }
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("== métricas geradas ==", file=sys.stderr)
    print(f"  hoje={out['hoje']} dados={B['date_min']}->{B['date_max']}", file=sys.stderr)
    print(f"  out={args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
